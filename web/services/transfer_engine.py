"""
Transfer Engine — Handles parsing template ITKs, matching CLR products
to templates, and writing transferred data into template files.
"""
import os
import re
import base64
import json
import urllib.parse
import openpyxl
from collections import defaultdict, Counter
from web.services.clr_parser import parse_clr


def parse_template_itks(filepath):
    """Parse a template file and extract its valid ITK aliases."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['Template']

    # Read headers from Row 4
    headers = []
    for cell in ws[4]:
        headers.append(str(cell.value).strip() if cell.value else None)

    # Extract settings from Row 1
    settings_cells = []
    for cell in ws[1]:
        if cell.value:
            settings_cells.append(str(cell.value))

    wb.close()

    # Decode attribute settings
    aliases = _decode_itk_aliases(settings_cells)

    # Derive product type from filename
    basename = os.path.basename(filepath)
    # Strip tmpl_ prefix if present (from our upload naming)
    if basename.startswith('tmpl_'):
        basename = basename[5:]
    product_type = os.path.splitext(basename)[0]

    # Build the base category slug from the product type name
    # e.g. HERBAL_SUPPLEMENT -> "herbal supplement" for browse-path matching
    base_category = product_type.replace('_', ' ').lower()
    # Also build a hyphenated version: "herbal-supplements" (with plural)
    base_slug = product_type.replace('_', '-').lower()
    # Try both singular and plural
    base_slugs = {base_slug, base_slug + 's'}
    # Also add space-separated versions for browse path matching
    base_browse_terms = {base_category, base_category + 's'}

    # Collect all specific alias slugs too
    alias_slugs = list(aliases.values())

    # Extract the parenthesized slug from each alias key for display
    # AND extract browse path segments for matching browse-path-only CLR ITKs
    slug_to_parenthesized = {}
    browse_path_segments = set()  # unique browse path segments from aliases
    for display_key, slug in aliases.items():
        # Extract the part in parentheses from the display key
        paren_match = re.search(r'\(([^)]+)\)\s*$', display_key)
        if paren_match:
            slug_to_parenthesized[slug] = paren_match.group(1)
            # Extract the browse path (everything before the parenthesized slug)
            browse_path = display_key[:paren_match.start()].strip()
        else:
            slug_to_parenthesized[slug] = slug
            browse_path = display_key.strip()

        # Store the full browse path and also its last few segments for matching
        # e.g. "Health & Household > ... > Blended Vitamin & Mineral Supplements"
        browse_path_lower = browse_path.lower().rstrip(' >')
        if browse_path_lower:
            browse_path_segments.add(browse_path_lower)
            # Also add just the last segment(s) for partial matching
            parts = [p.strip() for p in browse_path_lower.split('>')]
            if len(parts) >= 2:
                # Last 2 segments: "Herbal Supplements > Echinacea"
                browse_path_segments.add(' > '.join(parts[-2:]))
            if len(parts) >= 1:
                # Just the last segment: "Echinacea"
                browse_path_segments.add(parts[-1])

    return {
        'filepath': filepath,
        'filename': basename,
        'product_type': product_type,
        'headers': headers,
        'aliases': aliases,              # {display_key: slug}
        'slugs': alias_slugs,            # specific slugs from aliases
        'base_slugs': base_slugs,        # derived from filename
        'base_browse_terms': base_browse_terms,  # for browse path matching
        'browse_path_segments': browse_path_segments,  # from alias keys
        'slug_to_parenthesized': slug_to_parenthesized,
    }


def transfer_data(clr_path, template_paths):
    """Execute the full transfer pipeline."""
    # 1. Parse CLR
    clr_data = parse_clr(clr_path)

    # 2. Parse all templates
    templates = []
    for tp in template_paths:
        try:
            tmpl_info = parse_template_itks(tp)
            templates.append(tmpl_info)
        except Exception as e:
            continue

    if not templates:
        raise ValueError("No valid templates could be parsed.")

    # 3. Build family groups
    parent_map = {}  # sku -> product
    for p in clr_data['parents']:
        parent_map[p['sku']] = p

    # Also check if any "standalone" product is actually referenced as a parent
    all_by_sku = {p['sku']: p for p in clr_data['products']}

    child_families = defaultdict(list)  # parent_sku -> [children]
    for c in clr_data['children']:
        child_families[c['parent_sku']].append(c)

    # 4. Match products to templates and transfer
    summary_templates = []
    output_files = []
    all_matched_skus = set()
    total_transferred = 0
    total_flagged = 0

    for tmpl_idx, tmpl in enumerate(templates):
        result = _transfer_to_template(
            clr_data, tmpl, parent_map, child_families, all_by_sku
        )
        all_matched_skus.update(result['matched_skus'])
        total_transferred += result['total_rows']
        total_flagged += result['flagged_count']

        tmpl_summary = {
            'index': tmpl_idx,
            'filename': tmpl['filename'],
            'product_type': tmpl['product_type'],
            'families': result['family_count'],
            'standalone': result['standalone_count'],
            'flagged': result['flagged_count'],
            'total_rows': result['total_rows'],
            'matched_itks': result['matched_itks'],
            'flag_details': result['flag_details'],
        }
        summary_templates.append(tmpl_summary)

        output_files.append({
            'path': result['output_path'],
            'filename': tmpl['filename'],
        })

    # 5. Find unmatched products
    unmatched_products = []
    for p in clr_data['products']:
        if p['sku'] not in all_matched_skus:
            parentage = (p['parentage'] or '').lower()
            if 'parent' in parentage and 'child' not in parentage:
                ptype = 'Parent'
            elif 'child' in parentage:
                ptype = 'Child'
            else:
                ptype = 'Standalone'
            unmatched_products.append({
                'sku': p['sku'],
                'itk': p['itk'],
                'type': ptype,
            })

    summary = {
        'total_transferred': total_transferred,
        'templates_used': len([t for t in summary_templates if t['total_rows'] > 0]),
        'total_flagged': total_flagged,
        'unmatched': len(unmatched_products),
        'templates': summary_templates,
        'unmatched_products': unmatched_products,
    }

    return {
        'summary': summary,
        'output_files': output_files,
    }


def _product_matches_template(clr_itk_value, tmpl):
    """
    Check if a CLR product's ITK matches the template.

    Matching strategy (in order):
    1. Exact slug match: CLR ITK contains one of the template's alias slugs
       e.g. "echinacea-herbal-supplements" in "Echinacea (echinacea-herbal-supplements)"
    2. Browse path match: CLR ITK's browse path contains the base category
       e.g. "Herbal Supplements" in "Health & Household > ... > Herbal Supplements > Mushrooms"
    3. Slug-in-ITK match: CLR ITK contains the base slug from the filename
       e.g. "herbal-supplements" in "other-(herbal-supplements)"
    """
    if not clr_itk_value:
        return False
    itk_lower = str(clr_itk_value).lower().strip()

    # 1. Check specific alias slugs first
    for slug in tmpl['slugs']:
        if slug and slug.lower() in itk_lower:
            return True

    # 2. Check if CLR ITK's browse path matches any alias browse paths
    # e.g. CLR: "Health & Household > ... > Blended Vitamin & Mineral Supplements"
    # matches alias browse path from NUTRITIONAL_SUPPLEMENT template
    for segment in tmpl['browse_path_segments']:
        if segment in itk_lower:
            return True

    # 3. Check base browse terms (from filename)
    # e.g. "herbal supplements" matches "Health & Household > ... > Herbal Supplements > Mushrooms"
    for term in tmpl['base_browse_terms']:
        if term in itk_lower:
            return True

    # 4. Check base slugs (hyphenated form from filename)
    # e.g. "herbal-supplements" matches "other-(herbal-supplements)"
    for slug in tmpl['base_slugs']:
        if slug in itk_lower:
            return True

    return False


def _get_best_matching_slug(clr_itk_value, tmpl):
    """
    Return the best matching alias slug for a CLR ITK value.
    Prefers specific slug matches over base category matches.
    Returns (slug, is_specific) tuple.
    """
    if not clr_itk_value:
        return None, False
    itk_lower = str(clr_itk_value).lower().strip()

    # 1. Check specific alias slugs first
    for slug in tmpl['slugs']:
        if slug and slug.lower() in itk_lower:
            return slug, True

    # 2. For browse-path or base-slug matches, return None for slug
    #    (these don't have a specific alias, just the base category)
    return None, False


def _find_best_itk_display(clr_itk, tmpl):
    """
    Find the best ITK display value for a product.
    Uses the parenthesized slug format: (slug-name)

    Strategy:
    1. If CLR ITK matches a specific alias slug -> use that alias's parenthesized form
    2. If CLR ITK only matches the base category -> use the CLR's original ITK
       (since we can't determine the specific sub-category)
    """
    if not clr_itk:
        return None

    slug, is_specific = _get_best_matching_slug(clr_itk, tmpl)

    if is_specific and slug:
        # Use the parenthesized slug from the alias
        paren = tmpl['slug_to_parenthesized'].get(slug)
        if paren:
            return paren
        return slug

    # For base-category matches, keep the original CLR ITK value
    # The product matches the template but we don't have a specific slug for it
    return clr_itk


def _transfer_to_template(clr_data, tmpl, parent_map, child_families, all_by_sku):
    """Match CLR products to a single template and write the data."""
    # Find matching children
    matching_children = []
    for c in clr_data['children']:
        if _product_matches_template(c['itk'], tmpl):
            matching_children.append(c)

    # Find matching standalone
    matching_standalone = []
    for s in clr_data['standalone']:
        if _product_matches_template(s['itk'], tmpl):
            matching_standalone.append(s)

    # Build family groups from matching children
    matched_parent_skus = set()
    for c in matching_children:
        if c['parent_sku']:
            matched_parent_skus.add(c['parent_sku'])

    # Build ordered rows: families first, then standalone
    ordered_rows = []
    matched_skus = set()
    flag_details = []
    flagged_count = 0
    matched_itk_set = set()

    # Process families
    family_count = 0
    for parent_sku in matched_parent_skus:
        family_count += 1
        # Find the parent
        parent = parent_map.get(parent_sku) or all_by_sku.get(parent_sku)
        if parent:
            ordered_rows.append({
                'product': parent,
                'flag': None,
                'override_itk': True,
            })
            matched_skus.add(parent['sku'])

        # Get ALL children of this family
        all_family_children = child_families.get(parent_sku, [])
        # Determine majority ITK for the family
        family_itks = [c['itk'] for c in all_family_children if c['itk']]
        majority_itk = Counter(family_itks).most_common(1)[0][0] if family_itks else None

        for child in all_family_children:
            child_matches = _product_matches_template(child['itk'], tmpl)
            flag = None

            if not child_matches:
                if child['itk']:
                    flag = f"FLAGGED: Different ITK from family - original: {child['itk']}"
                else:
                    flag = "FLAGGED: Different ITK from family - original: (none)"
                flagged_count += 1
                flag_details.append({
                    'sku': child['sku'],
                    'reason': flag,
                })

            ordered_rows.append({
                'product': child,
                'flag': flag,
                'override_itk': child_matches,
            })
            matched_skus.add(child['sku'])

            if child_matches and child['itk']:
                display = _find_best_itk_display(child['itk'], tmpl)
                if display:
                    matched_itk_set.add(display)

        # Also track parent ITK
        if parent and majority_itk:
            display = _find_best_itk_display(majority_itk, tmpl)
            if display:
                matched_itk_set.add(display)

    # Process standalone
    for s in matching_standalone:
        ordered_rows.append({
            'product': s,
            'flag': None,
            'override_itk': True,
        })
        matched_skus.add(s['sku'])
        if s['itk']:
            display = _find_best_itk_display(s['itk'], tmpl)
            if display:
                matched_itk_set.add(display)

    # Now write to the template file
    output_path = _write_template(clr_data, tmpl, ordered_rows)

    # Build matched ITK display names for UI
    matched_itk_names = sorted(matched_itk_set)

    return {
        'output_path': output_path,
        'matched_skus': matched_skus,
        'total_rows': len(ordered_rows),
        'family_count': family_count,
        'standalone_count': len(matching_standalone),
        'flagged_count': flagged_count,
        'flag_details': flag_details,
        'matched_itks': matched_itk_names,
    }


def _write_template(clr_data, tmpl, ordered_rows):
    """Write product data into the template file."""
    wb = openpyxl.load_workbook(tmpl['filepath'], keep_vba=True)
    ws = wb['Template']

    # Build column mapping
    mapping = _build_ordinal_mapping(clr_data['headers'], tmpl['headers'])

    # Find key columns in template
    tmpl_pt_idx = None
    tmpl_itk_idx = None
    for i, h in enumerate(tmpl['headers']):
        if h and h.lower() == 'product type':
            tmpl_pt_idx = i
        if h and h.lower() == 'item type keyword':
            tmpl_itk_idx = i

    # Clear existing data from Row 5 onward
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(5, max(max_row + 1, 8)):
        for c in range(1, max_col + 2):
            ws.cell(row=r, column=c).value = None

    # Write attribute references to Row 5
    for clr_col, tmpl_col in mapping.items():
        if clr_col < len(clr_data['attr_refs']) and clr_data['attr_refs'][clr_col]:
            ws.cell(row=5, column=tmpl_col + 1, value=clr_data['attr_refs'][clr_col])

    # Row 6 stays empty

    # Add NOTES/FLAGS header
    notes_col = max_col + 1
    ws.cell(row=4, column=notes_col, value='NOTES/FLAGS')

    # Write product data starting at Row 7
    target_row = 7
    for entry in ordered_rows:
        product = entry['product']
        row_data = product['row_data']
        flag = entry['flag']
        override_itk = entry['override_itk']

        # Write mapped columns
        for clr_col, tmpl_col in mapping.items():
            if clr_col < len(row_data) and row_data[clr_col] is not None:
                ws.cell(row=target_row, column=tmpl_col + 1, value=row_data[clr_col])

        # Override Product Type
        if tmpl_pt_idx is not None:
            ws.cell(row=target_row, column=tmpl_pt_idx + 1, value=tmpl['product_type'])

        # Override ITK (only for matching products, not flagged ones)
        if tmpl_itk_idx is not None and override_itk:
            itk_display = _find_best_itk_display(product['itk'], tmpl)
            if itk_display:
                ws.cell(row=target_row, column=tmpl_itk_idx + 1, value=itk_display)

        # Write flag if present
        if flag:
            ws.cell(row=target_row, column=notes_col, value=flag)

        target_row += 1

    # Save
    output_path = tmpl['filepath'].replace('tmpl_', 'out_')
    wb.save(output_path)
    wb.close()

    return output_path


def _build_ordinal_mapping(clr_headers, template_headers):
    """Map CLR column indices to template column indices by header name with ordinal matching."""
    tmpl_occurrences = defaultdict(list)
    for i, h in enumerate(template_headers):
        if h is not None:
            tmpl_occurrences[h].append(i)

    clr_header_count = defaultdict(int)
    mapping = {}  # clr_col_index -> template_col_index

    for clr_idx, h in enumerate(clr_headers):
        if h is None:
            continue
        occurrence = clr_header_count[h]
        clr_header_count[h] += 1
        if h in tmpl_occurrences and occurrence < len(tmpl_occurrences[h]):
            mapping[clr_idx] = tmpl_occurrences[h][occurrence]

    return mapping


def _decode_itk_aliases(settings_cells):
    """Decode the ITK aliases from template Row 1 settings cells."""
    # 1. Combine all settings cells
    full_raw = ''
    for cell in settings_cells:
        match = re.match(r'settings\d*=(.*)', str(cell), re.DOTALL)
        if match:
            full_raw += match.group(1)

    # 2. URL-decode
    full_raw = urllib.parse.unquote(full_raw)

    # 3. Extract attributeSettings value
    idx = full_raw.find('attributeSettings=')
    if idx == -1:
        return {}

    rest = full_raw[idx + len('attributeSettings='):]
    next_key = re.search(r'&[a-zA-Z]+=', rest)
    b64_val = rest[:next_key.start()] if next_key else rest

    # 4. Fix base64 padding and decode
    while len(b64_val) % 4 != 0:
        b64_val += '='

    try:
        decoded = base64.b64decode(b64_val).decode('utf-8')
        data = json.loads(decoded)
    except Exception:
        return {}

    # 5. Find ITK aliases
    aliases = {}
    for item in data:
        if isinstance(item, dict) and 'item_type_keyword' in item.get('attribute', ''):
            aliases = item.get('aliases', {})
            break

    return aliases
