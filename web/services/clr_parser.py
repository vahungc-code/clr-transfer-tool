"""
CLR Parser — Reads Amazon Category Listing Report (.xlsm) files
and extracts product data, ITK values, and family groupings.
"""
import openpyxl
from collections import Counter


def parse_clr(filepath):
    """Parse a CLR file and return structured product data."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['Template']

    # Read headers from Row 4
    headers = []
    for cell in ws[4]:
        headers.append(str(cell.value).strip() if cell.value else None)

    # Read attribute references from Row 5
    attr_refs = []
    for cell in ws[5]:
        attr_refs.append(str(cell.value).strip() if cell.value else None)

    # Find key column indices
    col_map = _find_columns(headers)

    # Read product data starting from Row 6+
    # Find the actual first data row (skip attribute reference echoes)
    products = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        row_data = list(row)

        # Skip empty rows
        sku_idx = col_map.get('sku')
        if sku_idx is None or sku_idx >= len(row_data) or not row_data[sku_idx]:
            continue

        sku_val = str(row_data[sku_idx]).strip()

        # Skip attribute reference echo rows
        if sku_val.startswith('contribution_sku') or '#' in sku_val and '.value' in sku_val:
            continue

        # Skip rows where product_type looks like an attribute reference
        pt_idx = col_map.get('product_type')
        if pt_idx is not None and pt_idx < len(row_data) and row_data[pt_idx]:
            pt_val = str(row_data[pt_idx]).strip()
            if 'product_type#' in pt_val and '.value' in pt_val:
                continue

        product = {
            'row_data': row_data,
            'sku': sku_val,
            'product_type': _get_val(row_data, col_map.get('product_type')),
            'itk': _get_val(row_data, col_map.get('itk')),
            'parentage': _get_val(row_data, col_map.get('parentage')),
            'parent_sku': _get_val(row_data, col_map.get('parent_sku')),
        }
        products.append(product)

    wb.close()

    # Categorize products
    parents = []
    children = []
    standalone = []

    for p in products:
        parentage = (p['parentage'] or '').lower()
        if 'parent' in parentage and 'child' not in parentage:
            parents.append(p)
        elif 'child' in parentage and p['parent_sku']:
            children.append(p)
        else:
            standalone.append(p)

    return {
        'headers': headers,
        'attr_refs': attr_refs,
        'products': products,
        'parents': parents,
        'children': children,
        'standalone': standalone,
        'col_map': col_map,
        'total_products': len(products),
        'total_parents': len(parents),
        'total_children': len(children),
        'total_standalone': len(standalone),
    }


def extract_itk_summary(clr_data):
    """Extract unique ITK values with counts, sorted by count descending.

    For display, uses the parenthesized slug format:
    - If ITK already has a slug: "Blended Vitamin... (multiple-vitamin-mineral-combinations)"
      -> display as "(multiple-vitamin-mineral-combinations)"
    - If ITK is a browse path: "Health & Household > ... > Herbal Supplements > Mushrooms"
      -> derive slug from last segment -> display as "(mushrooms)"
    - Plain slug ITKs like "herbal-supplements" or "pet-herbal-supplements"
      -> display as "(herbal-supplements)" or "(pet-herbal-supplements)"
    """
    import re
    itk_counter = Counter()
    # Map raw ITK -> display name
    itk_display = {}

    for p in clr_data['products']:
        itk = p['itk']
        if itk:
            itk_counter[itk] += 1
            if itk not in itk_display:
                itk_display[itk] = _itk_to_slug_display(itk)
        else:
            itk_counter['(no ITK)'] += 1
            itk_display['(no ITK)'] = '(no ITK)'

    # Return tuples of (display_name, raw_itk, count)
    result = []
    for itk, count in sorted(itk_counter.items(), key=lambda x: x[1], reverse=True):
        result.append((itk_display.get(itk, itk), itk, count))

    return result


def _itk_to_slug_display(itk):
    """Convert any ITK format to a parenthesized slug display.

    Examples:
      "Blended Vitamin & Mineral Supplements (multiple-vitamin-mineral-combinations)"
        -> "(multiple-vitamin-mineral-combinations)"
      "Health & Household > Vitamins > Herbal Supplements > Mushrooms"
        -> "(mushrooms)"
      "Health & Household > Vitamins > Herbal Supplements"
        -> "(herbal-supplements)"
      "herbal-supplements"
        -> "(herbal-supplements)"
      "Herbal Supplements (herbal-supplements)"
        -> "(herbal-supplements)"
      "other-(herbal-supplements)"
        -> "(herbal-supplements)"
      "인삼-허브 보충제"  (non-ASCII)
        -> shown as-is
    """
    import re
    itk = itk.strip()

    # 1. Already has parenthesized slug at the end
    paren_match = re.search(r'\(([a-z0-9-]+)\)\s*$', itk)
    if paren_match:
        return f"({paren_match.group(1)})"

    # 2. Has parenthesized slug embedded (e.g. "other-(herbal-supplements)")
    embedded_match = re.search(r'\(([a-z0-9-]+)\)', itk)
    if embedded_match:
        return f"({embedded_match.group(1)})"

    # 3. Browse path format: "A > B > C > D" -> take last segment, slugify
    if ' > ' in itk:
        last_segment = itk.split('>')[-1].strip()
        slug = _slugify(last_segment)
        if slug:
            return f"({slug})"
        return itk

    # 4. Already looks like a slug (lowercase, hyphens, no spaces)
    if re.match(r'^[a-z0-9-]+$', itk):
        return f"({itk})"

    # 5. Plain text name -> slugify
    slug = _slugify(itk)
    if slug and slug != itk.lower():
        return f"({slug})"

    # 6. Fallback (non-ASCII, etc.)
    return itk


def _slugify(text):
    """Convert text to a URL-friendly slug.
    'Herbal Supplements' -> 'herbal-supplements'
    'Blended Vitamin & Mineral Supplements' -> 'blended-vitamin-mineral-supplements'
    """
    import re
    text = text.lower().strip()
    # Remove & and other special chars
    text = text.replace('&', '')
    # Replace spaces and special chars with hyphens
    text = re.sub(r'[^a-z0-9]+', '-', text)
    # Remove leading/trailing hyphens and collapse multiples
    text = re.sub(r'-+', '-', text).strip('-')
    return text


def _find_columns(headers):
    """Find indices of key columns by header name."""
    col_map = {}
    header_lower = [(h.lower() if h else '') for h in headers]

    mappings = {
        'sku': ['sku'],
        'product_type': ['product type'],
        'itk': ['item type keyword'],
        'parentage': ['parentage level', 'parentage'],
        'parent_sku': ['parent sku'],
    }

    for key, names in mappings.items():
        for i, h in enumerate(header_lower):
            if h in names:
                col_map[key] = i
                break

    return col_map


def _get_val(row_data, idx):
    """Safely get a string value from row data at the given index."""
    if idx is None or idx >= len(row_data) or row_data[idx] is None:
        return None
    val = str(row_data[idx]).strip()
    return val if val else None
